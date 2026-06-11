from __future__ import annotations

from io import BytesIO
import re
from typing import BinaryIO

import pandas as pd
from openpyxl import load_workbook

from bank_guarantee_tool.src.services.normalization_service import (
    normalize_date_column,
    normalize_guarantees,
    normalize_money_column,
    normalize_text_column,
)


COLUMN_ALIASES = {
    "guarantee_number": [
        "номер бг", "№ бг", "n бг", "гарантия", "номер гарантии", "номер банковской гарантии",
        "bg number", "guarantee number", "guarantee_number", "номер",
    ],
    "bank_name": ["банк", "банк гарант", "банк-гарант", "гарант", "bank", "bank_name", "issuing bank"],
    "supplier_name": [
        "поставщик", "бенефициар", "контрагент", "получатель", "beneficiary", "supplier",
        "supplier_name", "vendor", "vendor name",
    ],
    "supplier_inn": ["инн поставщика", "инн бенефициара", "инн", "inn", "supplier_inn", "tax id"],
    "legal_entity_name": [
        "юрлицо", "юр. лицо", "юридическое лицо", "принципал", "наша компания", "legal entity",
        "legal_entity_name", "principal", "company",
    ],
    "amount": ["сумма", "сумма бг", "лимит бг", "размер гарантии", "amount", "guarantee amount"],
    "start_date": ["дата начала", "начало", "действует с", "дата выдачи", "start date", "start_date", "valid from"],
    "end_date": ["дата окончания", "окончание", "срок до", "действует до", "end date", "end_date", "valid until"],
    "rate": ["ставка", "ставка %", "процент", "rate", "fee rate"],
    "actual_fee": ["комиссия", "факт комиссия", "стоимость", "actual_fee", "fee"],
    "status": ["статус", "состояние", "status", "state"],
    "total_limit": ["общий лимит", "лимит", "лимит банка", "total limit", "total_limit", "bank limit"],
    "reserved_limit": ["резерв", "зарезервировано", "reserved", "reserved_limit"],
    "limit_start_date": ["начало лимита", "дата начала лимита", "limit_start_date"],
    "limit_end_date": ["окончание лимита", "дата окончания лимита", "limit_end_date"],
    "limit_status": ["статус лимита", "limit_status"],
}

GUARANTEE_SCORE_COLUMNS = {"guarantee_number", "bank_name", "supplier_name", "amount", "end_date"}
LIMIT_SCORE_COLUMNS = {"bank_name", "legal_entity_name", "total_limit", "reserved_limit", "limit_status"}


def _clean_column_name(column: object) -> str:
    text = str(column).strip().lower().replace("ё", "е")
    text = text.replace("\n", " ").replace("\r", " ")
    text = re.sub(r"[№#]", " n ", text)
    text = re.sub(r"[^a-zа-я0-9%]+", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def _alias_to_canonical(value: object) -> str | None:
    cleaned = _clean_column_name(value)
    if not cleaned:
        return None
    for canonical, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            alias_cleaned = _clean_column_name(alias)
            if cleaned == alias_cleaned or alias_cleaned in cleaned:
                return canonical
    return None


def normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    rename_map = {}
    used_canonical: set[str] = set()
    for column in df.columns:
        canonical = _alias_to_canonical(column)
        if canonical and canonical not in used_canonical:
            rename_map[column] = canonical
            used_canonical.add(canonical)
    return df.rename(columns=rename_map)


def _row_match_score(values: list[object], expected_columns: set[str]) -> int:
    found = {_alias_to_canonical(value) for value in values}
    return len((found - {None}) & expected_columns)


def _detect_header_row(raw_df: pd.DataFrame, expected_columns: set[str]) -> int:
    if raw_df.empty:
        return 0
    max_rows = min(len(raw_df), 30)
    scored = [(_row_match_score(raw_df.iloc[index].tolist(), expected_columns), index) for index in range(max_rows)]
    best_score, best_index = max(scored, key=lambda item: item[0])
    return best_index if best_score > 0 else 0


def _stringify_header(value: object, fallback: str) -> str:
    if pd.isna(value) or str(value).strip() == "":
        return fallback
    return str(value).strip()


def _materialize_sheet(raw_df: pd.DataFrame, expected_columns: set[str]) -> tuple[pd.DataFrame, int, list[str]]:
    """Convert a raw sheet with arbitrary title rows into a table with detected headers."""
    if raw_df.empty:
        return pd.DataFrame(), 0, []
    header_row = _detect_header_row(raw_df, expected_columns)
    headers = [_stringify_header(value, f"unnamed_{i + 1}") for i, value in enumerate(raw_df.iloc[header_row].tolist())]
    table = raw_df.iloc[header_row + 1 :].copy()
    table.columns = headers
    table["source_excel_row"] = table.index + 1
    table = table.dropna(how="all")
    table = table.loc[:, ~table.columns.duplicated()].copy()
    return table.reset_index(drop=True), header_row + 1, headers


def load_excel_file(uploaded_file: BinaryIO) -> dict[str, pd.DataFrame]:
    """Read all workbook sheets as raw tables, including hidden sheets when the engine exposes them."""
    return pd.read_excel(uploaded_file, sheet_name=None, header=None, dtype=object)


def inspect_workbook(file_bytes: bytes) -> pd.DataFrame:
    """Return workbook sheet inventory: visibility, dimensions, merged ranges and row/column counts."""
    try:
        workbook = load_workbook(BytesIO(file_bytes), read_only=False, data_only=False)
    except Exception as exc:
        return pd.DataFrame([{"sheet_name": "workbook", "visibility": "unknown", "error": str(exc)}])
    rows: list[dict[str, object]] = []
    for index, worksheet in enumerate(workbook.worksheets, start=1):
        rows.append(
            {
                "sheet_index": index,
                "sheet_name": worksheet.title,
                "visibility": worksheet.sheet_state,
                "max_row": worksheet.max_row,
                "max_column": worksheet.max_column,
                "merged_ranges": ", ".join(str(item) for item in worksheet.merged_cells.ranges),
                "tables_count": len(worksheet.tables),
                "auto_filter": str(worksheet.auto_filter.ref or ""),
            }
        )
    return pd.DataFrame(rows)


def _score_sheet(raw_df: pd.DataFrame, expected_columns: set[str], preferred_names: set[str], sheet_name: str) -> int:
    name_bonus = 5 if _clean_column_name(sheet_name) in preferred_names else 0
    header_row = _detect_header_row(raw_df, expected_columns)
    header_score = _row_match_score(raw_df.iloc[header_row].tolist(), expected_columns) if not raw_df.empty else 0
    non_empty_bonus = 1 if not raw_df.dropna(how="all").empty else 0
    return name_bonus + header_score + non_empty_bonus


def detect_guarantee_sheet(sheets: dict[str, pd.DataFrame]) -> tuple[str, pd.DataFrame]:
    preferred_names = {"бг", "гарантии", "bank guarantees", "guarantees"}
    return max(
        sheets.items(),
        key=lambda item: _score_sheet(item[1], GUARANTEE_SCORE_COLUMNS, preferred_names, item[0]),
    )


def detect_limits_sheet(sheets: dict[str, pd.DataFrame]) -> tuple[str, pd.DataFrame] | None:
    preferred = {"лимиты", "лимиты банков", "bank limits", "limits"}
    if not sheets:
        return None
    sheet_name, raw_df = max(
        sheets.items(),
        key=lambda item: _score_sheet(item[1], LIMIT_SCORE_COLUMNS, preferred, item[0]),
    )
    if _score_sheet(raw_df, LIMIT_SCORE_COLUMNS, preferred, sheet_name) < 2:
        return None
    return sheet_name, raw_df


def prepare_guarantees_from_sheets(sheets: dict[str, pd.DataFrame]) -> tuple[str, pd.DataFrame]:
    sheet_name, raw_df = detect_guarantee_sheet(sheets)
    table, header_row, headers = _materialize_sheet(raw_df, GUARANTEE_SCORE_COLUMNS)
    normalized = normalize_columns(table)
    normalized["source_sheet"] = sheet_name
    normalized["source_header_row"] = header_row
    normalized["source_detected_headers"] = ", ".join(headers)
    guarantees = normalize_guarantees(normalized)
    key_fields = ["guarantee_number", "bank_name", "supplier_name", "amount", "end_date"]
    present_count = guarantees[key_fields].notna().sum(axis=1)
    present_count += (guarantees[["guarantee_number", "bank_name", "supplier_name"]] != "").sum(axis=1)
    guarantees["processing_status"] = "processed"
    guarantees.loc[present_count < 3, "processing_status"] = "needs_review"
    guarantees["processing_note"] = ""
    guarantees.loc[guarantees["processing_status"] == "needs_review", "processing_note"] = (
        "Строка прочитана, но распознано мало ключевых полей — проверьте заголовки и значения"
    )
    return sheet_name, guarantees


def prepare_limits_from_sheets(sheets: dict[str, pd.DataFrame]) -> tuple[str, pd.DataFrame] | None:
    detected = detect_limits_sheet(sheets)
    if detected is None:
        return None
    sheet_name, raw_df = detected
    table, header_row, headers = _materialize_sheet(raw_df, LIMIT_SCORE_COLUMNS)
    df = normalize_columns(table)
    df["source_sheet"] = sheet_name
    df["source_header_row"] = header_row
    df["source_detected_headers"] = ", ".join(headers)
    for column in ["bank_name", "legal_entity_name", "limit_status"]:
        if column not in df.columns:
            df[column] = ""
        df[column] = normalize_text_column(df[column])
    for column in ["total_limit", "reserved_limit"]:
        if column not in df.columns:
            df[column] = 0.0
        df[column] = normalize_money_column(df[column])
    for column in ["limit_start_date", "limit_end_date"]:
        if column in df.columns:
            df[column] = normalize_date_column(df[column])
    return sheet_name, df


def build_sheet_diagnostics(sheets: dict[str, pd.DataFrame]) -> pd.DataFrame:
    diagnostics: list[dict[str, object]] = []
    for sheet_name, raw_df in sheets.items():
        guarantee_header = _detect_header_row(raw_df, GUARANTEE_SCORE_COLUMNS)
        limit_header = _detect_header_row(raw_df, LIMIT_SCORE_COLUMNS)
        diagnostics.append(
            {
                "sheet_name": sheet_name,
                "raw_rows": len(raw_df),
                "raw_columns": len(raw_df.columns),
                "guarantee_score": _row_match_score(raw_df.iloc[guarantee_header].tolist(), GUARANTEE_SCORE_COLUMNS) if not raw_df.empty else 0,
                "guarantee_header_row": guarantee_header + 1,
                "limit_score": _row_match_score(raw_df.iloc[limit_header].tolist(), LIMIT_SCORE_COLUMNS) if not raw_df.empty else 0,
                "limit_header_row": limit_header + 1,
            }
        )
    return pd.DataFrame(diagnostics)
